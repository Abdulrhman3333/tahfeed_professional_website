from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from .forms import StudentRegistrationForm, StudentBulkUploadForm, MemorizationTemplateUploadForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from .models import Student, Attendance, TeacherAttendance, StageSupervisor, AcademicCalendar, ExamNomination, UserRole, TeacherPlanPreference, SmsTemplateSetting, MemorizationTemplateBundle, TeacherProfile
from django.utils import timezone
from django.contrib.auth.models import User
from django.db.models import Count, Q, Max
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from uuid import uuid4
import json
import re
from urllib import request as urlrequest
from urllib.error import HTTPError, URLError
from urllib.parse import quote, unquote


class TeacherLoginView(LoginView):
    """تسجيل دخول المعلمين مع خيار تذكر اسم المستخدم والاستمرار في تسجيل الدخول."""
    template_name = 'login.html'

    def get_initial(self):
        initial = super().get_initial()
        remembered_username = self.request.COOKIES.get('remembered_username')
        if remembered_username:
            try:
                remembered_username = unquote(remembered_username)
            except Exception:
                remembered_username = ''
        if remembered_username:
            initial['username'] = remembered_username
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        remembered_username = self.request.COOKIES.get('remembered_username', '')
        if remembered_username:
            try:
                remembered_username = unquote(remembered_username)
            except Exception:
                remembered_username = ''
        context['remembered_username'] = remembered_username
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        remember_me = self.request.POST.get('remember_me') == 'on'
        username = form.cleaned_data.get('username', '')

        # Keep user logged in for a long period by default after first login.
        self.request.session.set_expiry(60 * 60 * 24 * 30)

        if remember_me:
            # Cookie values must be ASCII-safe; encode to avoid crashes with Arabic usernames.
            safe_username = quote(username, safe='')
            response.set_cookie(
                'remembered_username',
                safe_username,
                max_age=60 * 60 * 24 * 30,
                httponly=False,
                samesite='Lax'
            )
        else:
            response.delete_cookie('remembered_username')

        return response

# دالة مساعدة للتحقق من صلاحية مشرف المرحلة
def is_stage_supervisor(user):
    """تحقق إذا كان المستخدم مشرف مرحلة"""
    return hasattr(user, 'stage_supervisor') or user.is_superuser or user_has_role(user, 'supervisor')


def user_has_role(user, role_code):
    """تحقق من العضوية"""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return UserRole.objects.filter(user=user, role__code=role_code).exists()


def can_access_preparer_attendance(user):
    """السماح للمحضّر أو المشرف (بأنواعه) بالدخول لصفحات متابعة/تحضير الحضور العامة."""
    return user_has_role(user, 'preparer') or is_stage_supervisor(user)


ARABIC_WEEKDAY_NAMES = {
    0: 'الأحد',
    1: 'الاثنين',
    2: 'الثلاثاء',
    3: 'الأربعاء',
    4: 'الخميس',
    5: 'الجمعة',
    6: 'السبت',
}


def get_arabic_weekday_name(target_date):
    return ARABIC_WEEKDAY_NAMES.get(target_date.weekday(), 'الأحد')


def normalize_saudi_phone(phone):
    phone = phone or ''
    digits = ''.join([ch for ch in phone if ch.isdigit()])
    if digits.startswith('966'):
        return digits
    if digits.startswith('0'):
        return '966' + digits[1:]
    if digits.startswith('5'):
        return '966' + digits
    return digits


def normalize_excel_value(value):
    if value is None:
        return ''
    return str(value).strip()


def normalize_arabic_text(value):
    text = normalize_excel_value(value)
    replacements = {
        'أ': 'ا',
        'إ': 'ا',
        'آ': 'ا',
        'ى': 'ي',
        'ة': 'ه',
    }
    for old_char, new_char in replacements.items():
        text = text.replace(old_char, new_char)
    return ' '.join(text.split()).lower()


def map_grade_from_excel(raw_grade):
    text_value = normalize_excel_value(raw_grade)
    if not text_value:
        return ''

    # Accept internal code directly (e.g. 3_pri).
    valid_codes = {choice[0] for choice in Student.GRADE_CHOICES}
    if text_value in valid_codes:
        return text_value

    normalized_text = normalize_arabic_text(text_value)

    # Values that mean "not selected yet" are accepted as empty grade.
    empty_markers = {
        normalize_arabic_text('لم يتم التحديد بعد'),
        normalize_arabic_text('غير محدد'),
        normalize_arabic_text('غير محدده'),
        normalize_arabic_text('لا يوجد'),
    }
    if normalized_text in empty_markers:
        return ''

    # Accept official labels and common writing variants.
    label_map = {
        normalize_arabic_text('أول ابتدائي'): '1_pri',
        normalize_arabic_text('اول ابتدائي'): '1_pri',
        normalize_arabic_text('اولى ابتدائي'): '1_pri',
        normalize_arabic_text('اولي ابتدائي'): '1_pri',
        normalize_arabic_text('ثاني ابتدائي'): '2_pri',
        normalize_arabic_text('ثانيه ابتدائي'): '2_pri',
        normalize_arabic_text('ثالث ابتدائي'): '3_pri',
        normalize_arabic_text('ثالثه ابتدائي'): '3_pri',
        normalize_arabic_text('رابع ابتدائي'): '4_pri',
        normalize_arabic_text('رابعه ابتدائي'): '4_pri',
        normalize_arabic_text('خامس ابتدائي'): '5_pri',
        normalize_arabic_text('خامسه ابتدائي'): '5_pri',
        normalize_arabic_text('سادس ابتدائي'): '6_pri',
        normalize_arabic_text('سادسه ابتدائي'): '6_pri',
        normalize_arabic_text('أول متوسط'): '1_med',
        normalize_arabic_text('اول متوسط'): '1_med',
        normalize_arabic_text('اولى متوسط'): '1_med',
        normalize_arabic_text('اولي متوسط'): '1_med',
        normalize_arabic_text('ثاني متوسط'): '2_med',
        normalize_arabic_text('ثالث متوسط'): '3_med',
        normalize_arabic_text('أول ثانوي'): '1_sec',
        normalize_arabic_text('اول ثانوي'): '1_sec',
        normalize_arabic_text('ثاني ثانوي'): '2_sec',
        normalize_arabic_text('ثالث ثانوي'): '3_sec',
        normalize_arabic_text('جامعي'): 'uni',
    }

    return label_map.get(normalized_text, '')


def parse_excel_birth_date(value):
    if value in (None, ''):
        return ''

    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)):
        try:
            base_date = datetime(1899, 12, 30)
            parsed = base_date + timedelta(days=float(value))
            return parsed.strftime('%Y-%m-%d')
        except Exception:
            return str(value)

    text = str(value).strip()
    digit_map = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')
    text = text.translate(digit_map)
    return text


def read_excel_sheet_as_dicts(uploaded_file):
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_excel_value(item) for item in rows[0]]
    data_rows = []

    for row_idx, row_values in enumerate(rows[1:], start=2):
        row_data = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row_values[col_idx] if col_idx < len(row_values) else None
        data_rows.append({'excel_row': row_idx, 'data': row_data})

    return data_rows


def sheet_to_dict_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_excel_value(item) for item in rows[0]]
    result = []

    for row_values in rows[1:]:
        row_data = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row_values[col_idx] if col_idx < len(row_values) else None
        result.append(row_data)

    return result


def parse_memorization_bundle_from_workbook(workbook):
    metadata_sheet = workbook['metadata'] if 'metadata' in workbook.sheetnames else None
    if metadata_sheet is None and 'meta' in workbook.sheetnames:
        metadata_sheet = workbook['meta']

    if metadata_sheet is not None:
        metadata_rows = sheet_to_dict_rows(metadata_sheet)
        metadata_map = {
            normalize_excel_value(row.get('key')): row.get('value')
            for row in metadata_rows
            if normalize_excel_value(row.get('key'))
        }

        template_map = {}
        for key, value in metadata_map.items():
            match = re.match(r'^template\.(\d+)\.(value|dataKey|title|description|sheet)$', key)
            if not match:
                continue
            idx = int(match.group(1))
            field = match.group(2)
            if idx not in template_map:
                template_map[idx] = {}
            template_map[idx][field] = normalize_excel_value(value)

        template_defs = []
        memorization_data = {}

        for idx in sorted(template_map.keys()):
            item = template_map[idx]
            value = normalize_excel_value(item.get('value'))
            data_key = normalize_excel_value(item.get('dataKey'))
            title = normalize_excel_value(item.get('title'))
            description = normalize_excel_value(item.get('description'))
            sheet_name = normalize_excel_value(item.get('sheet'))

            if not value or not data_key or not title or not sheet_name:
                continue
            if sheet_name not in workbook.sheetnames:
                continue

            data_rows = sheet_to_dict_rows(workbook[sheet_name])
            entries = [
                normalize_excel_value(row.get('entryText') or row.get('entry') or row.get('value'))
                for row in data_rows
            ]
            entries = [entry for entry in entries if entry]
            if not entries:
                continue

            memorization_data[data_key] = {
                'data': entries,
                'index': {},
            }
            template_defs.append({
                'value': value,
                'dataKey': data_key,
                'title': title,
                'description': description,
            })

        if template_defs and memorization_data:
            return {
                'template_definitions': template_defs,
                'memorization_data': memorization_data,
            }

    # Legacy format fallback: templates + entries + indexes sheets.
    if 'entries' not in workbook.sheetnames:
        raise ValueError('الملف لا يحتوي على ورقة metadata أو entries بالصيغة المعتمدة.')

    template_rows = sheet_to_dict_rows(workbook['templates']) if 'templates' in workbook.sheetnames else []
    data_rows = sheet_to_dict_rows(workbook['entries'])
    index_rows = sheet_to_dict_rows(workbook['indexes']) if 'indexes' in workbook.sheetnames else []

    memorization_data = {}
    for row in data_rows:
        data_key = normalize_excel_value(row.get('dataKey'))
        entry_text = normalize_excel_value(row.get('entryText'))
        entry_index_raw = row.get('entryIndex')
        if not data_key or not entry_text:
            continue

        try:
            entry_index = int(entry_index_raw)
        except (TypeError, ValueError):
            continue

        if data_key not in memorization_data:
            memorization_data[data_key] = {'data': [], 'index': {}}

        data_list = memorization_data[data_key]['data']
        while len(data_list) <= entry_index:
            data_list.append('')
        data_list[entry_index] = entry_text

    for payload in memorization_data.values():
        payload['data'] = [entry for entry in payload['data'] if entry]

    for row in index_rows:
        data_key = normalize_excel_value(row.get('dataKey'))
        index_key = normalize_excel_value(row.get('indexKey'))
        if not data_key or not index_key or data_key not in memorization_data:
            continue

        try:
            index_value = int(row.get('indexValue'))
        except (TypeError, ValueError):
            continue

        if 0 <= index_value < len(memorization_data[data_key]['data']):
            memorization_data[data_key]['index'][index_key] = index_value

    template_defs = []
    if template_rows:
        for row in template_rows:
            value = normalize_excel_value(row.get('value'))
            data_key = normalize_excel_value(row.get('dataKey'))
            title = normalize_excel_value(row.get('title'))
            description = normalize_excel_value(row.get('description'))
            if not value or not data_key or not title or data_key not in memorization_data:
                continue
            template_defs.append({
                'value': value,
                'dataKey': data_key,
                'title': title,
                'description': description,
            })

    if not template_defs:
        for idx, data_key in enumerate(memorization_data.keys(), start=1):
            template_defs.append({
                'value': f'custom_{idx}',
                'dataKey': data_key,
                'title': f'خطة {data_key}',
                'description': 'خطة مستوردة',
            })

    if not memorization_data:
        raise ValueError('لا توجد بيانات حفظ صالحة داخل الملف المرفوع.')

    return {
        'template_definitions': template_defs,
        'memorization_data': memorization_data,
    }


def build_status_export_rows(target_date, status_code):
    attendances = Attendance.objects.filter(date=target_date, status=status_code).select_related('student')
    rows = []

    for attendance in attendances:
        student = attendance.student
        teacher = student.teacher
        full_name = student.full_name or ''
        first_name = full_name.split()[0] if full_name.split() else ''
        total_days = Attendance.objects.filter(student=student, status=status_code).count()
        teacher_name = ''
        teacher_phone = ''

        if teacher:
            teacher_name = teacher.get_full_name() or teacher.username or ''
            profile = getattr(teacher, 'teacher_profile', None)
            teacher_phone = normalize_saudi_phone(getattr(profile, 'phone', '')) if profile else ''

        rows.append({
            'phone_number': normalize_saudi_phone(student.parent_phone),
            'first_name': first_name,
            'full_name': full_name,
            'teacher_name': teacher_name,
            'teacher_phone': teacher_phone,
            'total_days': total_days,
            'today_date': target_date,
            'today_day_name': get_arabic_weekday_name(target_date),
        })

    return rows


def export_status_rows_to_excel(rows, status_label, target_date):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = status_label
    worksheet.sheet_view.rightToLeft = True

    headers = ['رقم الجوال', 'الاسم الأول', 'الاسم الكامل', f'إجمالي أيام {status_label}', 'تاريخ اليوم', 'اسم اليوم']
    worksheet.append(headers)

    for row in rows:
        worksheet.append([
            row['phone_number'],
            row['first_name'],
            row['full_name'],
            row['total_days'],
            row['today_date'].strftime('%Y-%m-%d'),
            row['today_day_name'],
        ])

    filename = f'preparer_contacts_{status_label}_{target_date.strftime("%Y-%m-%d")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def render_message_template(template_text, row_data):
    """Simple token renderer for templates like: Hello {{{first_name}}}."""
    if not template_text:
        return ''

    placeholders = {
        'first_name': row_data.get('first_name', ''),
        'full_name': row_data.get('full_name', ''),
        'phone_number': row_data.get('phone_number', ''),
        'teacher_name': row_data.get('teacher_name', ''),
        'teacher_phone': row_data.get('teacher_phone', ''),
        'total_days': row_data.get('total_days', ''),
        'part_number': row_data.get('part_number', ''),
        'today_date': row_data.get('today_date', ''),
        'today_day_name': row_data.get('today_day_name', ''),
        # Arabic aliases for template variables.
        'الاسم_الأول': row_data.get('first_name', ''),
        'الاسم_الكامل': row_data.get('full_name', ''),
        'رقم_الجوال': row_data.get('phone_number', ''),
        'اسم_المعلم': row_data.get('teacher_name', ''),
        'جوال_المعلم': row_data.get('teacher_phone', ''),
        'رقم_جوال_المعلم': row_data.get('teacher_phone', ''),
        'المجموع': row_data.get('total_days', ''),
        'رقم_الجزء': row_data.get('part_number', ''),
        'التاريخ': row_data.get('today_date', ''),
        'اليوم': row_data.get('today_day_name', ''),
    }

    rendered = template_text
    for key, value in placeholders.items():
        rendered = rendered.replace(f'{{{{{{{key}}}}}}}', str(value or ''))
    return rendered.strip()


def send_sms_via_api(phone_number, message_text):
    """Send one SMS using a configurable JSON API endpoint."""
    api_url = getattr(settings, 'SMS_API_URL', '').strip()
    api_key = getattr(settings, 'SMS_API_KEY', '').strip()

    if not api_url:
        return False, 'SMS API URL is not configured.'
    if not api_key:
        return False, 'SMS API key is not configured.'

    phone_field = getattr(settings, 'SMS_PHONE_FIELD', 'to')
    message_field = getattr(settings, 'SMS_MESSAGE_FIELD', 'message')
    sender_field = getattr(settings, 'SMS_SENDER_FIELD', 'sender')
    sender_id = getattr(settings, 'SMS_SENDER_ID', '').strip()

    phone_is_array = bool(getattr(settings, 'SMS_PHONE_IS_ARRAY', False))

    payload = {
        phone_field: [phone_number] if phone_is_array else phone_number,
        message_field: message_text,
    }

    if sender_id:
        payload[sender_field] = sender_id
    elif sender_field == 'src':
        return False, 'SMS_SENDER_ID is not configured (required field: src).'

    auth_header = getattr(settings, 'SMS_AUTH_HEADER', 'Authorization')
    auth_scheme = getattr(settings, 'SMS_AUTH_SCHEME', 'Bearer').strip()
    auth_value = f'{auth_scheme} {api_key}' if auth_scheme else api_key

    body = json.dumps(payload).encode('utf-8')
    req = urlrequest.Request(
        api_url,
        data=body,
        headers={
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            auth_header: auth_value,
        },
        method='POST',
    )

    try:
        timeout = int(getattr(settings, 'SMS_API_TIMEOUT', 15))
        with urlrequest.urlopen(req, timeout=timeout) as response:
            status_code = getattr(response, 'status', 200)
            return 200 <= status_code < 300, f'HTTP {status_code}'
    except HTTPError as exc:
        return False, f'HTTP {exc.code}'
    except URLError as exc:
        return False, f'URL error: {exc.reason}'
    except Exception as exc:
        return False, str(exc)


def send_batch_messages(rows, template_text):
    """Send templated messages to each row phone and return a summary."""
    success_count = 0
    failed_count = 0
    failures = []

    for row in rows:
        phone_number = row.get('phone_number')
        if not phone_number:
            failed_count += 1
            failures.append('رقم هاتف فارغ')
            continue

        message_text = render_message_template(template_text, row)
        if not message_text:
            failed_count += 1
            failures.append(f'{phone_number}: الرسالة فارغة')
            continue

        ok, info = send_sms_via_api(phone_number, message_text)
        if ok:
            success_count += 1
        else:
            failed_count += 1
            failures.append(f'{phone_number}: {info}')

    return {
        'success_count': success_count,
        'failed_count': failed_count,
        'failures': failures[:5],
    }

@login_required
def pending_students(request):
    """صفحة الطلاب المنتظرين - للمشرفين فقط"""
    # التحقق من الصلاحيات
    if not is_stage_supervisor(request.user):
        return redirect('home')
    
    # جلب الطلاب حسب مرحلة المشرف
    if request.user.is_superuser:
        # المدير يرى جميع الطلاب
        students = Student.objects.filter(status='منتظر').order_by('full_name')
    elif hasattr(request.user, 'stage_supervisor'):
        # مشرف المرحلة يرى فقط طلاب مرحلته
        supervisor = request.user.stage_supervisor
        students = Student.objects.filter(status='منتظر', educational_stage=supervisor.stage).order_by('full_name')
    elif user_has_role(request.user, 'supervisor'):
        # المشرف العام يرى جميع الطلاب
        students = Student.objects.filter(status='منتظر').order_by('full_name')
    else:
        students = Student.objects.none()
    
    # جلب جميع المعلمين لقائمة التعيين (المعلمين الحاليين + الموظفين)
    teachers = User.objects.filter(
        Q(student__isnull=False) | Q(is_staff=True)
    ).distinct().order_by('username')
    
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        teacher_id = request.POST.get('teacher_id')
        student = get_object_or_404(Student, id=student_id)
        
        # تعيين المعلم إذا تم اختياره
        if teacher_id:
            teacher = get_object_or_404(User, id=teacher_id)
            student.teacher = teacher
        
        student.status = 'منتظم'
        student.save()
        return redirect('pending_students')

    return render(request, 'pending.html', {'students': students, 'teachers': teachers})

@login_required
def stage_students_data(request):
    """صفحة عرض بيانات طلاب المرحلة - للمشرفين فقط"""
    # التحقق من الصلاحيات
    if not is_stage_supervisor(request.user):
        return redirect('home')
    
    # جلب الطلاب حسب مرحلة المشرف
    if request.user.is_superuser:
        # المدير يرى جميع الطلاب المنتظمين
        students = Student.objects.filter(status='منتظم')
    elif hasattr(request.user, 'stage_supervisor'):
        # مشرف المرحلة يرى فقط طلاب مرحلته المنتظمين
        supervisor = request.user.stage_supervisor
        students = Student.objects.filter(status='منتظم', educational_stage=supervisor.stage)
    elif user_has_role(request.user, 'supervisor'):
        # المشرف العام يرى جميع الطلاب المنتظمين
        students = Student.objects.filter(status='منتظم')
    else:
        students = Student.objects.none()
    
    today = timezone.now().date()
    # النطاق الافتراضي: من الأحد إلى الخميس من الأسبوع الحالي.
    sunday_offset = (today.weekday() + 1) % 7
    default_start_date = today - timedelta(days=sunday_offset)
    default_end_date = default_start_date + timedelta(days=4)

    # فلترة التاريخ لاحتساب الغياب ضمن نطاق محدد
    start_date_str = (request.GET.get('start_date') or '').strip()
    end_date_str = (request.GET.get('end_date') or '').strip()

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else default_start_date
    except ValueError:
        start_date = default_start_date

    try:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else default_end_date
    except ValueError:
        end_date = default_end_date

    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    # نحتاج قائمة فعلية لإضافة العد ثم الترتيب حسب الأعلى غياباً
    students = list(students.select_related('teacher').order_by('teacher__username', 'full_name'))

    for student in students:
        attendance_qs = Attendance.objects.filter(student=student, status='غائب')
        if start_date:
            attendance_qs = attendance_qs.filter(date__gte=start_date)
        if end_date:
            attendance_qs = attendance_qs.filter(date__lte=end_date)
        student.absent_count_in_range = attendance_qs.count()

    students.sort(key=lambda s: (-getattr(s, 'absent_count_in_range', 0), s.full_name or ''))
    
    # جلب جميع المعلمين لقائمة التعيين
    teachers = User.objects.filter(
        Q(student__isnull=False) | Q(is_staff=True)
    ).distinct().order_by('username')
    
    # معالجة طلب تحديث المعلم
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        teacher_id = request.POST.get('teacher_id')
        
        try:
            student = get_object_or_404(Student, id=student_id)
            
            # تعيين المعلم الجديد إذا تم اختياره
            if teacher_id:
                teacher = get_object_or_404(User, id=teacher_id)
                student.teacher = teacher
                student.save()
        except:
            pass

        query = []
        if start_date:
            query.append(f"start_date={start_date.strftime('%Y-%m-%d')}")
        if end_date:
            query.append(f"end_date={end_date.strftime('%Y-%m-%d')}")

        if query:
            return redirect(f"{reverse('stage_students_data')}?{'&'.join(query)}")
        return redirect('stage_students_data')
    
    # الحصول على المرحلة الحالية
    current_stage = None
    if hasattr(request.user, 'stage_supervisor'):
        current_stage = request.user.stage_supervisor.stage
    
    return render(request, 'stage_students_data.html', {
        'students': students,
        'teachers': teachers,
        'current_stage': current_stage,
        'start_date': start_date,
        'end_date': end_date,
        'today': today,
        'default_start_date': default_start_date,
        'default_end_date': default_end_date,
    })

@login_required
def take_attendance(request):
    """التحضير اليومي للمعلم"""
    # جلب طلاب هذا المعلم فقط والذين حالتهم "منتظم"
    students = Student.objects.filter(teacher=request.user, status='منتظم').order_by('full_name')
    
    # الحصول على التاريخ الحالي
    today = timezone.now().date()

    selected_date_str = request.POST.get('attendance_date') or request.GET.get('date') or str(today)
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        selected_date = today
    
    # تحديد اليوم بالعربي
    weekday_map = {
        6: 'الأحد',      # Sunday
        0: 'الاثنين',    # Monday
        1: 'الثلاثاء',   # Tuesday
        2: 'الأربعاء',   # Wednesday
        3: 'الخميس',    # Thursday
    }
    current_weekday = weekday_map.get(selected_date.weekday(), 'الأحد')
    
    # الحصول على رقم الأسبوع من التقويم
    current_week = AcademicCalendar.get_week_from_date(selected_date)
    
    success = False
    if request.method == 'POST':
        posted_date_str = request.POST.get('attendance_date', str(selected_date))
        try:
            posted_date = datetime.strptime(posted_date_str, "%Y-%m-%d").date()
        except ValueError:
            posted_date = selected_date

        selected_weekday = weekday_map.get(posted_date.weekday(), 'الأحد')
        selected_week = AcademicCalendar.get_week_from_date(posted_date)
        
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status:
                Attendance.objects.update_or_create(
                    student=student,
                    date=posted_date,
                    defaults={
                        'status': status,
                        'weekday': selected_weekday,
                        'week_number': selected_week
                    }
                )
        # بدل redirect، نعيد تحميل البيانات وإظهار رسالة نجاح
        success = True
        selected_date = posted_date
        current_weekday = selected_weekday
        current_week = selected_week

    existing_attendance = Attendance.objects.filter(student__in=students, date=selected_date)
    attendance_map = {item.student_id: item.status for item in existing_attendance}
    
    # إضافة حالة لكل طالب (حاضر افتراضياً إذا لم يكن له سجل)
    students_with_status = []
    for student in students:
        status = attendance_map.get(student.id, 'حاضر')
        students_with_status.append({
            'student': student,
            'status': status
        })
    
    context = {
        'students_with_status': students_with_status,
        'current_weekday': current_weekday,
        'current_week': current_week,
        'selected_date': selected_date,
        'today': today,
        'attendance_map': attendance_map,
        'success': success,
    }
    
    return render(request, 'attendance.html', context)

def success_view(request):
    return render(request, 'success.html')

def attendance_success_view(request):
    return render(request, 'attendance_success.html')

def welcome_view(request):
    return render(request, 'welcome.html')

def parent_inquiry(request):
    error_message = None
    if request.method == 'POST':
        parent_phone = request.POST.get('parent_phone', '').strip()
        if not parent_phone:
            error_message = "يرجى إدخال رقم الجوال"
        else:
            students = Student.objects.filter(parent_phone=parent_phone).select_related('teacher').order_by('full_name')

            for student in students:
                attendances = Attendance.objects.filter(student=student).order_by('-date')
                student.present_count = attendances.filter(status='حاضر').count()
                student.absent_count = attendances.filter(status='غائب').count()
                student.absent_excused_count = attendances.filter(status='غياب بعذر').count()
                student.excused_count = attendances.filter(status='مستأذن').count()
                student.late_count = attendances.filter(status='متأخر').count()
                student.recent_attendance = attendances[:10]

                teacher_phone = None
                if student.teacher and hasattr(student.teacher, 'teacher_profile'):
                    teacher_phone = student.teacher.teacher_profile.phone
                student.teacher_phone = teacher_phone

            return render(
                request,
                'parent_inquiry_results.html',
                {
                    'students': students,
                    'parent_phone': parent_phone,
                }
            )

    return render(request, 'parent_inquiry.html', {'error_message': error_message})

# دالة للتحقق من أن المستخدم ليس معلماً أو مشرفاً (للصفحة العامة فقط)
def home(request):
    """الصفحة الرئيسية - نموذج التسجيل العام"""
    # السماح للمشرفين والمدير بالوصول لصفحة التسجيل
    if request.user.is_authenticated:
        # المدير والمشرفون يمكنهم الوصول
        if not (is_stage_supervisor(request.user) or request.user.is_superuser):
            # المعلمون العاديون يتم توجيههم
            if Student.objects.filter(teacher=request.user).exists():
                return redirect('teacher_dashboard')
            elif request.user.is_staff:
                return redirect('admin:index')
    
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('success')
    else:
        form = StudentRegistrationForm()
    
    return render(request, 'home.html', {'form': form})


@login_required
def bulk_students_upload(request):
    if not (request.user.is_superuser or user_has_role(request.user, 'manager')):
        return redirect('home')

    form = StudentBulkUploadForm(request.POST or None, request.FILES or None)
    result = None

    if request.method == 'POST' and form.is_valid():
        uploaded_file = form.cleaned_data['excel_file']
        last_part_choices = {choice[0] for choice in Student._meta.get_field('last_tested_part').choices}
        status_default = Student._meta.get_field('status').default

        created_count = 0
        skipped_count = 0
        row_errors = []

        try:
            rows = read_excel_sheet_as_dicts(uploaded_file)
        except Exception:
            form.add_error('excel_file', 'تعذر قراءة الملف. تأكد أن الملف بصيغة Excel صحيحة (.xlsx).')
            rows = []

        for row in rows:
            row_number = row['excel_row']
            data = row['data']

            full_name = normalize_excel_value(data.get('full_name'))
            identity_number = normalize_excel_value(data.get('identity_number'))

            if not full_name and not identity_number:
                skipped_count += 1
                continue

            if not identity_number:
                identity_number = f'GEN_{uuid4().hex[:12]}'

            if Student.objects.filter(identity_number=identity_number).exists():
                skipped_count += 1
                row_errors.append(f'السطر {row_number}: رقم الهوية {identity_number} موجود مسبقاً.')
                continue

            grade_input = data.get('grade')
            grade = map_grade_from_excel(grade_input)
            if normalize_excel_value(grade_input) and not grade:
                skipped_count += 1
                row_errors.append(f'السطر {row_number}: الصف الدراسي "{normalize_excel_value(grade_input)}" غير صحيح.')
                continue

            last_tested_part = normalize_excel_value(data.get('last_tested_part')) or '0'
            if last_tested_part not in last_part_choices:
                last_tested_part = '0'

            birth_date = parse_excel_birth_date(data.get('birth_date'))

            Student.objects.create(
                full_name=full_name,
                student_phone=normalize_excel_value(data.get('student_phone')),
                parent_phone=normalize_excel_value(data.get('parent_phone')),
                identity_number=identity_number,
                jamiaa_id=normalize_excel_value(data.get('jamiaa_id')),
                parent_identity=normalize_excel_value(data.get('parent_identity')),
                grade=grade,
                birth_date=birth_date,
                last_tested_part=last_tested_part,
                previous_center=normalize_excel_value(data.get('previous_center')),
                neighborhood=normalize_excel_value(data.get('neighborhood')),
                status=status_default,
            )
            created_count += 1

        if created_count > 0:
            form = StudentBulkUploadForm()

        result = {
            'created_count': created_count,
            'skipped_count': skipped_count,
            'row_errors': row_errors,
        }

    return render(request, 'bulk_students_upload.html', {
        'form': form,
        'result': result,
    })


@login_required
@user_passes_test(lambda user: user.is_superuser)
def superuser_template_upload(request):
    form = MemorizationTemplateUploadForm(request.POST or None, request.FILES or None)
    result = None

    if request.method == 'POST' and form.is_valid():
        uploaded_file = form.cleaned_data['excel_file']

        try:
            workbook = load_workbook(uploaded_file, data_only=True)
            parsed = parse_memorization_bundle_from_workbook(workbook)

            MemorizationTemplateBundle.objects.filter(is_active=True).update(is_active=False)
            saved_bundle = MemorizationTemplateBundle.objects.create(
                name='قوالب حفظ مرفوعة من المدير',
                source_filename=uploaded_file.name,
                template_definitions=parsed['template_definitions'],
                memorization_data=parsed['memorization_data'],
                is_active=True,
                uploaded_by=request.user,
            )

            result = {
                'saved': True,
                'bundle_id': saved_bundle.id,
                'template_count': len(parsed['template_definitions']),
                'source_filename': uploaded_file.name,
            }
            form = MemorizationTemplateUploadForm()
        except ValueError as exc:
            form.add_error('excel_file', str(exc))
        except Exception:
            form.add_error('excel_file', 'تعذر معالجة الملف. تأكد أن الملف مطابق لصيغة قوالب الحفظ المعتمدة.')

    active_bundle = MemorizationTemplateBundle.objects.filter(is_active=True).first()

    return render(request, 'superuser_template_upload.html', {
        'form': form,
        'result': result,
        'active_bundle': active_bundle,
    })


@login_required
def active_memorization_bundle_api(request):
    bundle = MemorizationTemplateBundle.objects.filter(is_active=True).first()
    if not bundle:
        return JsonResponse({'available': False})

    return JsonResponse({
        'available': True,
        'templateDefinitions': bundle.template_definitions,
        'memorizationData': bundle.memorization_data,
        'uploadedAt': bundle.created_at.isoformat(),
        'sourceFilename': bundle.source_filename,
    })

@login_required
def teacher_dashboard(request):
    """لوحة تحكم المعلم - عرض الإحصائيات"""
    # جلب طلاب المعلم
    students = Student.objects.filter(teacher=request.user, status='منتظم').order_by('full_name')

    # فلترة التاريخ لعرض الإحصائيات في نطاق محدد
    today = timezone.now().date()
    start_date_str = (request.GET.get('start_date') or '').strip()
    end_date_str = (request.GET.get('end_date') or '').strip()

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else None
    except ValueError:
        start_date = None

    try:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else None
    except ValueError:
        end_date = None

    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date
    
    # إحصائيات عامة
    total_students = students.count()
    
    stats_rows = []
    for student in students:
        attendances = Attendance.objects.filter(student=student)
        if start_date:
            attendances = attendances.filter(date__gte=start_date)
        if end_date:
            attendances = attendances.filter(date__lte=end_date)
        
        total_days = attendances.count()
        present = attendances.filter(status='حاضر').count()
        absent = attendances.filter(status='غائب').count()
        absent_excused = attendances.filter(status='غياب بعذر').count()
        excused = attendances.filter(status='مستأذن').count()
        late = attendances.filter(status='متأخر').count()
        
        # عدد الأيام غير الحاضر (غياب + استئذان + تأخير)
        non_present_count = absent + absent_excused + excused + late
        
        # جميع الحضور مع الحالة
        all_attendance = list(attendances.values('date', 'status').order_by('-date'))
        
        stats_rows.append((student.id, {
            'student': student,
            'total_days': total_days,
            'present': present,
            'absent': absent,
            'absent_excused': absent_excused,
            'excused': excused,
            'late': late,
            'non_present_count': non_present_count,
            'all_attendance': all_attendance
        }))

    stats_rows.sort(key=lambda item: (-item[1]['absent'], item[1]['student'].full_name or ''))
    attendance_stats = {student_id: stats for student_id, stats in stats_rows}
    
    context = {
        'total_students': total_students,
        'attendance_stats': attendance_stats,
        'start_date': start_date,
        'end_date': end_date,
        'today': today,
    }
    
    return render(request, 'teacher_dashboard.html', context)


@login_required
def teacher_students_data(request):
    """عرض بيانات طلاب المعلم الحالي."""
    preference, _ = TeacherPlanPreference.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        mem_plan = request.POST.get('mem_plan', preference.mem_plan)
        big_review_pages = request.POST.get('big_review_pages', preference.big_review_pages)

        if mem_plan not in {'none', '2', '1', '0.5', '0.25'}:
            mem_plan = '1'

        try:
            big_review_pages = int(big_review_pages)
            if big_review_pages < 1:
                big_review_pages = 1
        except (TypeError, ValueError):
            big_review_pages = preference.big_review_pages

        preference.mem_plan = mem_plan
        preference.big_review_pages = big_review_pages
        preference.save()

        return redirect('teacher_students_data')

    students = Student.objects.filter(teacher=request.user, status='منتظم').order_by('full_name')

    students_data = []
    for student in students:
        try:
            parts_count = int(student.last_tested_part)
        except (TypeError, ValueError):
            parts_count = 0

        students_data.append({
            'student': student,
            'parts_count': parts_count,
        })

    return render(
        request,
        'teacher_students_data.html',
        {
            'students_data': students_data,
            'preference': preference,
        },
    )


@login_required
def delete_teacher_student(request, student_id):
    """حذف طالب من قائمة المعلم الحالي."""
    student = get_object_or_404(Student, id=student_id, teacher=request.user, status='منتظم')

    if request.method == 'POST':
        student.delete()

    return redirect('teacher_students_data')


@login_required
def teacher_student_plan(request, student_id):
    """صفحة إنشاء خطة لطالب واحد باستخدام تفضيلات المعلم الافتراضية."""
    student = get_object_or_404(
        Student,
        id=student_id,
        teacher=request.user,
        status='منتظم',
    )
    preference, _ = TeacherPlanPreference.objects.get_or_create(user=request.user)

    student_payload = {
        'id': str(student.student_unique_id or student.id),
        'name': student.full_name,
        'db_id': student.id,
    }

    return render(
        request,
        'teacher_student_plan.html',
        {
            'student': student,
            'student_payload': student_payload,
            'preference': preference,
        },
    )


@login_required
def update_attendance(request):
    """تحديث حالة الحضور من لوحة تحكم المعلم"""
    if request.method == 'POST':
        from datetime import datetime
        
        # جلب جميع البيانات المرسلة
        for key, value in request.POST.items():
            if key.startswith('status_'):
                # استخراج التاريخ والـ student_id
                # الصيغة: status_YYYY-MM-DD_student_id
                try:
                    # إزالة البادئة 'status_'
                    rest = key.replace('status_', '')
                    # فصل بآخر underscore لأن التاريخ قد يحتوي على شرطات
                    last_underscore = rest.rfind('_')
                    date_str = rest[:last_underscore]
                    student_id = rest[last_underscore+1:]
                    
                    attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    student = Student.objects.get(id=student_id, teacher=request.user)
                    
                    # تحديث حالة الحضور
                    attendance = Attendance.objects.filter(
                        student=student,
                        date=attendance_date
                    ).first()
                    
                    if attendance:
                        attendance.status = value
                        attendance.save()
                except (Student.DoesNotExist, ValueError):
                    pass
    
    return redirect('teacher_dashboard')

@login_required
def nominate_for_exam(request):
    """صفحة ترشيح الطلاب للاختبار الداخلي"""
    # جلب طلاب المعلم
    students = Student.objects.filter(teacher=request.user, status='منتظم').order_by('full_name')

    existing_nominations = ExamNomination.objects.filter(teacher=request.user)
    nomination_map = {item.student_id: item.id for item in existing_nominations}
    
    if request.method == 'POST':
        for student in students:
            teacher_grade = request.POST.get(f'teacher_grade_{student.id}')
            if teacher_grade:
                ExamNomination.objects.create(
                    student=student,
                    teacher=request.user,
                    last_tested_part=student.last_tested_part,
                    teacher_grade=teacher_grade
                )
        return redirect('teacher_nominations')
    
    return render(request, 'nominate_exam.html', {
        'students': students,
        'nomination_map': nomination_map,
    })


@login_required
def delete_nomination(request, nomination_id):
    """حذف ترشيح طالب من الاختبار"""
    nomination = get_object_or_404(ExamNomination, id=nomination_id)

    can_delete = (
        nomination.teacher_id == request.user.id
        or user_has_role(request.user, 'examiner')
        or user_has_role(request.user, 'manager')
        or request.user.is_superuser
    )
    if not can_delete:
        return redirect('home')

    if request.method == 'POST':
        nomination.delete()
    next_url = request.POST.get('next') or 'teacher_nominations'
    return redirect(next_url)


@login_required
def delete_pending_student(request, student_id):
    """حذف طالب من قائمة المنتظرين"""
    if not is_stage_supervisor(request.user):
        return redirect('home')

    student = get_object_or_404(Student, id=student_id, status='منتظر')

    if not request.user.is_superuser and hasattr(request.user, 'stage_supervisor'):
        supervisor = request.user.stage_supervisor
        if student.educational_stage != supervisor.stage:
            return redirect('pending_students')

    if request.method == 'POST':
        student.delete()
    return redirect('pending_students')


@login_required
def teacher_nominations(request):
    """عرض الطلاب المرشحين من قبل المعلم"""
    nominations = ExamNomination.objects.filter(teacher=request.user).select_related('student').order_by('student__full_name', '-id')

    nominations_with_next = []
    for nomination in nominations:
        nominations_with_next.append({
            'nomination': nomination,
            'next_part': nomination.get_next_part()
        })

    return render(request, 'teacher_nominations.html', {'nominations': nominations_with_next})

@login_required
def nominated_students(request):
    """صفحة المرشحين للاختبار"""
    if not user_has_role(request.user, 'examiner'):
        return redirect('home')

    nominations = ExamNomination.objects.filter(internal_passed=False).select_related('student', 'teacher').order_by('student__full_name', '-id')
    
    if request.method == 'POST':
        for nomination in nominations:
            internal_grade = request.POST.get(f'internal_grade_{nomination.id}')
            if internal_grade:
                nomination.internal_grade = internal_grade
                teacher_ok = nomination.teacher_grade is not None and float(nomination.teacher_grade) >= 85
                internal_ok = float(internal_grade) >= 85
                nomination.internal_passed = teacher_ok and internal_ok
                nomination.save()
        return redirect('nominated_students')
    
    # إضافة الجزء التالي لكل ترشيح
    nominations_with_next = []
    for nomination in nominations:
        nominations_with_next.append({
            'nomination': nomination,
            'next_part': nomination.get_next_part()
        })

    mae_source = ExamNomination.objects.filter(
        teacher_grade__isnull=False,
        internal_grade__isnull=False,
    ).select_related('teacher')

    mae_map = {}
    for nomination in mae_source:
        teacher = nomination.teacher
        if teacher.id not in mae_map:
            mae_map[teacher.id] = {
                'teacher_name': teacher.get_full_name() or teacher.username,
                'sum_abs_error': 0.0,
                'count': 0,
            }

        abs_error = abs(float(nomination.teacher_grade) - float(nomination.internal_grade))
        mae_map[teacher.id]['sum_abs_error'] += abs_error
        mae_map[teacher.id]['count'] += 1

    teacher_mae_rows = []
    for item in mae_map.values():
        count = item['count']
        mae_value = item['sum_abs_error'] / count if count else 0.0
        teacher_mae_rows.append({
            'teacher_name': item['teacher_name'],
            'samples_count': count,
            'mae': mae_value,
        })

    teacher_mae_rows.sort(key=lambda row: row['mae'])

    return render(request, 'nominated_students.html', {
        'nominations': nominations_with_next,
        'teacher_mae_rows': teacher_mae_rows,
    })


@login_required
def association_candidates(request):
    """مرشحو اختبار الجمعية"""
    if not user_has_role(request.user, 'examiner'):
        return redirect('home')

    nominations = ExamNomination.objects.filter(internal_passed=True, association_tested=False).select_related('student', 'exam_halaqa_teacher').order_by('student__full_name', '-id')

    halaqa_profiles = TeacherProfile.objects.filter(
        class_name__isnull=False
    ).exclude(
        class_name__exact=''
    ).select_related('user').order_by('class_name', 'user__username')
    halaqa_teacher_ids = {profile.user_id for profile in halaqa_profiles}

    default_sms_template = (
        "السلام عليكم ورحمة الله وبركاته\n\n"
        "نذكركم باختبار ابنكم {{{الاسم_الأول}}} بعدد أجزاء {{{رقم_الجزء}}} في جمعية خيركم\n"
        "يوم ( يتم كتابة اليوم والتاريخ بشكل يدوي هنا )\n\n"
        "يرجى الحضور الى جامع الحمودي.\n"
        "الساعة (يتم كتابة الوقت بشكل يدوي هنا)\n"
        "من جهة بوابة دورات المياه.\n\n"
        "وذلك للذهاب مع زملائه الى مقر الجمعية لآداء الاختبار.\n\n"
        "وجزاكم الله خيراً\n"
        "نرجو الالتزام بالموعد المحدد"
    )

    saved_sms_template = SmsTemplateSetting.objects.filter(
        user=request.user,
        section='association_exam'
    ).values_list('template_text', flat=True).first()

    association_sms_template = saved_sms_template or default_sms_template
    posted_sms_template = request.POST.get('association_sms_template')
    if posted_sms_template is not None:
        association_sms_template = posted_sms_template

    sms_feedback = None
    template_save_feedback = None

    if request.method == 'POST':
        form_action = request.POST.get('form_action')

        if form_action == 'save_association_sms_template':
            template_text = (request.POST.get('association_sms_template') or '').strip() or default_sms_template
            SmsTemplateSetting.objects.update_or_create(
                user=request.user,
                section='association_exam',
                defaults={'template_text': template_text},
            )
            association_sms_template = template_text
            template_save_feedback = 'تم حفظ قالب رسالة اختبار الجمعية بنجاح.'

        elif form_action == 'send_association_sms':
            selected_ids_csv = (request.POST.get('selected_nomination_ids_csv') or '').strip()
            selected_ids = []
            if selected_ids_csv:
                for raw_id in selected_ids_csv.split(','):
                    raw_id = raw_id.strip()
                    if not raw_id:
                        continue
                    try:
                        selected_ids.append(int(raw_id))
                    except ValueError:
                        continue

            selected_nominations = nominations.filter(id__in=selected_ids) if selected_ids else nominations.none()

            nomination_rows = []
            for nomination in selected_nominations:
                student = nomination.student
                phone_number = normalize_saudi_phone(student.parent_phone)
                first_name = (student.full_name or '').strip().split()[0] if (student.full_name or '').strip() else ''
                nomination_rows.append({
                    'first_name': first_name,
                    'full_name': student.full_name,
                    'phone_number': phone_number,
                    'part_number': nomination.get_next_part(),
                })

            if not nomination_rows:
                sms_feedback = {
                    'success_count': 0,
                    'failed_count': 0,
                    'failures': ['لم يتم تحديد أي طالب لإرسال الرسالة.'],
                }
            else:
                batch_result = send_batch_messages(nomination_rows, association_sms_template)
                sms_feedback = {
                    'success_count': batch_result['success_count'],
                    'failed_count': batch_result['failed_count'],
                    'failures': batch_result['failures'],
                }

        else:
            for nomination in nominations:
                exam_halaqa_teacher_raw = (request.POST.get(f'exam_halaqa_teacher_{nomination.id}') or '').strip()
                if exam_halaqa_teacher_raw:
                    try:
                        selected_halaqa_teacher_id = int(exam_halaqa_teacher_raw)
                    except ValueError:
                        selected_halaqa_teacher_id = None
                    if selected_halaqa_teacher_id not in halaqa_teacher_ids:
                        selected_halaqa_teacher_id = None
                else:
                    selected_halaqa_teacher_id = None

                nomination.exam_halaqa_teacher_id = selected_halaqa_teacher_id
                association_grade = request.POST.get(f'association_grade_{nomination.id}')
                if association_grade:
                    nomination.association_grade = association_grade
                    nomination.association_tested = True
                nomination.save()
            return redirect('association_candidates')

    nominations_with_next = []
    for nomination in nominations:
        nominations_with_next.append({
            'nomination': nomination,
            'next_part': nomination.get_next_part()
        })

    return render(request, 'association_candidates.html', {
        'nominations': nominations_with_next,
        'halaqa_profiles': halaqa_profiles,
        'association_sms_template': association_sms_template,
        'sms_feedback': sms_feedback,
        'template_save_feedback': template_save_feedback,
    })


@login_required
def association_results(request):
    """نتائج اختبار الجمعية"""
    if not (user_has_role(request.user, 'examiner') or user_has_role(request.user, 'manager')):
        return redirect('home')

    nominations = ExamNomination.objects.filter(association_tested=True).select_related('student').order_by('student__full_name', '-id')
    return render(request, 'association_results.html', {'nominations': nominations})


@login_required
def preparer_attendance_summary(request):
    """متابعة تحضير المعلمين لليوم"""
    if not can_access_preparer_attendance(request.user):
        return redirect('home')

    target_date = timezone.now().date()
    teachers = User.objects.filter(student__isnull=False).distinct().order_by('username')
    pending = []
    completed = []

    for teacher in teachers:
        student_count = Student.objects.filter(teacher=teacher, status='منتظم').count()
        if student_count == 0:
            continue
        attendance_qs = Attendance.objects.filter(student__teacher=teacher, date=target_date)
        recorded_count = attendance_qs.count()
        display_name = teacher.get_full_name() or teacher.username

        if attendance_qs.count() >= student_count:
            last_time = attendance_qs.aggregate(last_time=Max('created_at'))['last_time']
            completed.append({
                'teacher': teacher,
                'display_name': display_name,
                'student_count': student_count,
                'recorded_count': recorded_count,
                'last_time': last_time,
            })
        else:
            pending.append({
                'teacher': teacher,
                'display_name': display_name,
                'student_count': student_count,
                'recorded_count': recorded_count,
            })

    context = {
        'target_date': target_date,
        'pending': pending,
        'completed': completed,
    }
    return render(request, 'preparer_attendance_summary.html', context)


@login_required
def preparer_take_attendance(request):
    """تحضير المعلمين - للمحضّر فقط"""
    if not can_access_preparer_attendance(request.user):
        return redirect('home')

    # teachers = User.objects.filter(student__isnull=False).distinct().order_by('username')
    teachers = User.objects.distinct().order_by('username')

    today = timezone.now().date()
    selected_date_str = request.POST.get('attendance_date') or request.GET.get('date') or str(today)
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        selected_date = today

    weekday_map = {
        6: 'الأحد',
        0: 'الاثنين',
        1: 'الثلاثاء',
        2: 'الأربعاء',
        3: 'الخميس',
    }
    current_weekday = weekday_map.get(selected_date.weekday(), 'الأحد')
    current_week = AcademicCalendar.get_week_from_date(selected_date)

    success = False
    if request.method == 'POST':
        posted_date_str = request.POST.get('attendance_date', str(selected_date))
        try:
            posted_date = datetime.strptime(posted_date_str, "%Y-%m-%d").date()
        except ValueError:
            posted_date = selected_date

        selected_weekday = weekday_map.get(posted_date.weekday(), 'الأحد')
        selected_week = AcademicCalendar.get_week_from_date(posted_date)

        for teacher in teachers:
            status = request.POST.get(f'status_{teacher.id}')
            if status:
                TeacherAttendance.objects.update_or_create(
                    teacher=teacher,
                    date=posted_date,
                    defaults={
                        'status': status,
                        'weekday': selected_weekday,
                        'week_number': selected_week
                    }
                )

        success = True
        selected_date = posted_date
        current_weekday = selected_weekday
        current_week = selected_week

    existing_attendance = TeacherAttendance.objects.filter(teacher__in=teachers, date=selected_date)
    attendance_map = {item.teacher_id: item.status for item in existing_attendance}

    teachers_with_status = []
    for teacher in teachers:
        status = attendance_map.get(teacher.id, 'حاضر')
        display_name = teacher.get_full_name() or teacher.username
        teachers_with_status.append({
            'teacher': teacher,
            'display_name': display_name,
            'status': status,
        })

    teachers_with_status.sort(key=lambda item: item['display_name'])

    context = {
        'teachers_with_status': teachers_with_status,
        'current_weekday': current_weekday,
        'current_week': current_week,
        'selected_date': selected_date,
        'today': today,
        'success': success,
    }

    return render(request, 'preparer_take_attendance.html', context)


@login_required
def preparer_take_students_attendance(request):
    """تحضير طلاب أي معلم - للمحضّر أو المدير"""
    if not can_access_preparer_attendance(request.user):
        return redirect('home')

    teachers = User.objects.filter(student__status='منتظم').distinct().order_by('username')

    today = timezone.now().date()
    selected_date_str = request.POST.get('attendance_date') or request.GET.get('date') or str(today)
    selected_teacher_id = request.POST.get('teacher_id') or request.GET.get('teacher_id')

    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        selected_date = today

    weekday_map = {
        6: 'الأحد',
        0: 'الاثنين',
        1: 'الثلاثاء',
        2: 'الأربعاء',
        3: 'الخميس',
    }
    current_weekday = weekday_map.get(selected_date.weekday(), 'الأحد')
    current_week = AcademicCalendar.get_week_from_date(selected_date)

    selected_teacher = None
    students = Student.objects.none()
    if selected_teacher_id:
        try:
            selected_teacher = teachers.get(id=selected_teacher_id)
            students = Student.objects.filter(teacher=selected_teacher, status='منتظم').order_by('full_name')
        except (User.DoesNotExist, ValueError, TypeError):
            selected_teacher = None
            students = Student.objects.none()

    success = False
    error_message = None

    if request.method == 'POST':
        posted_date_str = request.POST.get('attendance_date', str(selected_date))
        try:
            posted_date = datetime.strptime(posted_date_str, "%Y-%m-%d").date()
        except ValueError:
            posted_date = selected_date

        selected_weekday = weekday_map.get(posted_date.weekday(), 'الأحد')
        selected_week = AcademicCalendar.get_week_from_date(posted_date)

        if not selected_teacher:
            error_message = 'يرجى اختيار المعلم أولاً.'
        else:
            for student in students:
                status = request.POST.get(f'status_{student.id}')
                if status:
                    Attendance.objects.update_or_create(
                        student=student,
                        date=posted_date,
                        defaults={
                            'status': status,
                            'weekday': selected_weekday,
                            'week_number': selected_week,
                        }
                    )

            success = True
            selected_date = posted_date
            current_weekday = selected_weekday
            current_week = selected_week

    existing_attendance = Attendance.objects.filter(student__in=students, date=selected_date)
    attendance_map = {item.student_id: item.status for item in existing_attendance}

    students_with_status = []
    for student in students:
        status = attendance_map.get(student.id, 'حاضر')
        students_with_status.append({
            'student': student,
            'status': status,
        })

    teachers_with_names = []
    for teacher in teachers:
        teachers_with_names.append({
            'teacher': teacher,
            'display_name': teacher.get_full_name() or teacher.username,
        })

    teachers_with_names.sort(key=lambda item: item['display_name'])

    context = {
        'teachers': teachers_with_names,
        'selected_teacher': selected_teacher,
        'students_with_status': students_with_status,
        'selected_date': selected_date,
        'today': today,
        'current_weekday': current_weekday,
        'current_week': current_week,
        'success': success,
        'error_message': error_message,
    }
    return render(request, 'preparer_take_students_attendance.html', context)


@login_required
def preparer_absent_contacts(request):
    """أرقام أولياء الأمور للغائبين وإحصائية الغياب"""
    if not user_has_role(request.user, 'preparer'):
        return redirect('home')

    target_date_str = request.GET.get('date') or request.POST.get('date') or str(timezone.now().date())
    download_type = request.GET.get('download')
    try:
        target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
    except ValueError:
        target_date = timezone.now().date()

    default_templates = {
        'absent': 'ولي الأمر المكرم، نفيدكم بغياب الابن {{{الاسم_الأول}}} اليوم {{{اليوم}}} {{{التاريخ}}}. مجموع أيام غيابه حتى الآن: ({{{المجموع}}}) أيام. للاستفسار: المعلم {{{اسم_المعلم}}} - {{{جوال_المعلم}}}.\nإدارة جامع الحمودي',
        'absent_excused': 'ولي الأمر المكرم، نفيدكم بغياب الابن {{{الاسم_الأول}}} بعذر اليوم {{{اليوم}}} {{{التاريخ}}}. مجموع مرات الغياب بعذر حتى الآن: ({{{المجموع}}}) يوم. للاستفسار: المعلم {{{اسم_المعلم}}} - {{{جوال_المعلم}}}.\nإدارة جامع الحمودي',
        'late': 'ولي الأمر المكرم، نحيطكم علماً بتأخر الابن {{{الاسم_الأول}}} عن حلقة اليوم {{{اليوم}}} {{{التاريخ}}}. مجموع مرات التأخير: ({{{المجموع}}}). للاستفسار: المعلم {{{اسم_المعلم}}} - {{{جوال_المعلم}}}.\nإدارة جامع الحمودي',
        'excused': 'ولي الأمر المكرم، نود إحاطتكم بانصراف الابن {{{الاسم_الأول}}} وخروجه قبل نهاية وقت الحلقة اليوم {{{اليوم}}}. مجموع مرات الانصراف: ({{{المجموع}}}). للاستفسار: المعلم {{{اسم_المعلم}}} - {{{جوال_المعلم}}}.\nإدارة جامع الحمودي',
    }

    saved_templates = {
        row.section: row.template_text
        for row in SmsTemplateSetting.objects.filter(user=request.user)
    }

    current_templates = {
        'absent': saved_templates.get('absent', default_templates['absent']),
        'absent_excused': saved_templates.get('absent_excused', default_templates['absent_excused']),
        'late': saved_templates.get('late', default_templates['late']),
        'excused': saved_templates.get('excused', default_templates['excused']),
    }

    posted_templates = {
        'absent': request.POST.get('absent_sms_template'),
        'absent_excused': request.POST.get('absent_excused_sms_template'),
        'late': request.POST.get('late_sms_template'),
        'excused': request.POST.get('excused_sms_template'),
    }
    for key, value in posted_templates.items():
        if value is not None:
            current_templates[key] = value

    template_save_feedback = None
    if request.method == 'POST':
        reset_student_id = request.POST.get('reset_student_id')
        if reset_student_id:
            student = get_object_or_404(Student, id=reset_student_id)
            student.absence_reset_at = timezone.now().date()
            student.save()
            return redirect(f"{reverse('preparer_absent_contacts')}?date={target_date}")

        if request.POST.get('save_templates') == '1':
            template_field_map = {
                'absent': 'absent_sms_template',
                'absent_excused': 'absent_excused_sms_template',
                'late': 'late_sms_template',
                'excused': 'excused_sms_template',
            }
            saved_count = 0
            for section, field_name in template_field_map.items():
                if field_name not in request.POST:
                    continue
                template_text = (request.POST.get(field_name) or '').strip()
                if not template_text:
                    template_text = default_templates[section]

                SmsTemplateSetting.objects.update_or_create(
                    user=request.user,
                    section=section,
                    defaults={'template_text': template_text},
                )
                current_templates[section] = template_text
                saved_count += 1

            template_save_feedback = f'تم حفظ القالب بنجاح ({saved_count}).'

    def collect_parent_phones_by_status(status_code):
        status_attendance = Attendance.objects.filter(date=target_date, status=status_code).select_related('student')
        phones = []
        for attendance in status_attendance:
            formatted = normalize_saudi_phone(attendance.student.parent_phone)
            if formatted:
                phones.append(formatted)

        unique_phones = sorted(set(phones))
        return {
            'phones': unique_phones,
            'phones_text': "\n".join(unique_phones),
            'phones_count': len(unique_phones),
        }

    absent_contacts = collect_parent_phones_by_status('غائب')
    absent_excused_contacts = collect_parent_phones_by_status('غياب بعذر')
    late_contacts = collect_parent_phones_by_status('متأخر')
    excused_contacts = collect_parent_phones_by_status('مستأذن')

    status_rows = {
        'absent': build_status_export_rows(target_date, 'غائب'),
        'absent_excused': build_status_export_rows(target_date, 'غياب بعذر'),
        'late': build_status_export_rows(target_date, 'متأخر'),
        'excused': build_status_export_rows(target_date, 'مستأذن'),
    }

    sms_feedback = None
    if request.method == 'POST':
        sms_action = request.POST.get('sms_action')
        if sms_action in status_rows and request.POST.get('save_templates') != '1':
            batch_result = send_batch_messages(status_rows[sms_action], current_templates[sms_action])
            sms_feedback = {
                'section': sms_action,
                'success_count': batch_result['success_count'],
                'failed_count': batch_result['failed_count'],
                'failures': batch_result['failures'],
            }

    if download_type in {'absent', 'absent_excused', 'late', 'excused'}:
        status_map = {
            'absent': ('غائب', 'غياب'),
            'absent_excused': ('غياب بعذر', 'غياب_بعذر'),
            'late': ('متأخر', 'تأخر'),
            'excused': ('مستأذن', 'انصراف'),
        }
        status_code, status_label = status_map[download_type]
        rows = build_status_export_rows(target_date, status_code)
        return export_status_rows_to_excel(rows, status_label, target_date)

    at_risk = []
    students = Student.objects.filter(status='منتظم').order_by('full_name')
    for student in students:
        total_absences = Attendance.objects.filter(
            student=student,
            status='غائب'
        ).count()

        if student.absence_reset_at:
            temp_count = Attendance.objects.filter(
                student=student,
                status='غائب',
                created_at__date__gt=student.absence_reset_at
            ).count()
        else:
            temp_count = Attendance.objects.filter(
                student=student,
                status='غائب'
            ).count()
        if temp_count >= 5:
            at_risk.append({
                'student': student,
                'temp_count': temp_count,
                'total_absences': total_absences,
            })

    context = {
        'target_date': target_date,
        'absent_phones_text': absent_contacts['phones_text'],
        'absent_phones_count': absent_contacts['phones_count'],
        'absent_excused_phones_text': absent_excused_contacts['phones_text'],
        'absent_excused_phones_count': absent_excused_contacts['phones_count'],
        'late_phones_text': late_contacts['phones_text'],
        'late_phones_count': late_contacts['phones_count'],
        'excused_phones_text': excused_contacts['phones_text'],
        'excused_phones_count': excused_contacts['phones_count'],
        'absent_sms_template': current_templates['absent'],
        'absent_excused_sms_template': current_templates['absent_excused'],
        'late_sms_template': current_templates['late'],
        'excused_sms_template': current_templates['excused'],
        'sms_feedback': sms_feedback,
        'template_save_feedback': template_save_feedback,
        'at_risk': at_risk,
    }
    return render(request, 'preparer_absent_contacts.html', context)

# 1. للمدير فقط: لا يدخل إلا من لديه صلاحيات Superuser
def is_admin(user):
    return user.is_superuser or user_has_role(user, 'manager')

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    # كود لوحة التحكم الشاملة
    return render(request, 'admin_view.html')

@login_required
@user_passes_test(is_admin)
def admin_statistics(request):
    """صفحة الإحصائيات الشاملة للمدير"""
    from django.db.models import Count, Q
    
    # الحصول على جميع المراحل الموجودة
    STAGES = [
        ('مبكرة', 'مبكرة'),
        ('عليا', 'عليا'),
        ('متوسط', 'متوسط'),
        ('ثانوي', 'ثانوي'),
        ('جامعي', 'جامعي'),
    ]
    
    # إحصائيات عامة
    total_students = Student.objects.filter(status='منتظم').count()
    total_tested = ExamNomination.objects.filter(internal_passed=True).count()
    
    # إجمالي الغياب اليومي
    today = timezone.now().date()
    total_absent_today = Attendance.objects.filter(date=today, status='غائب').count()
    total_teacher_absent_today = TeacherAttendance.objects.filter(date=today, status='غائب').count()
    
    # إحصائيات حسب المرحلة
    stage_stats = []
    for stage_code, stage_name in STAGES:
        stage_students = Student.objects.filter(status='منتظم', educational_stage=stage_code)
        stage_tested = ExamNomination.objects.filter(
            student__educational_stage=stage_code,
            internal_passed=True
        ).count()
        
        # الغياب اليومي في هذه المرحلة
        stage_absent_today = Attendance.objects.filter(
            student__educational_stage=stage_code,
            date=today,
            status='غائب'
        ).count()
        
        if stage_students.count() > 0:
            stage_stats.append({
                'stage': stage_name,
                'total_students': stage_students.count(),
                'tested': stage_tested,
                'absent_today': stage_absent_today,
            })
    
    # إحصائيات حسب المعلم
    teachers = User.objects.filter(
        student__status='منتظم'
    ).distinct().order_by('username')
    
    teacher_stats = []
    for teacher in teachers:
        teacher_students = Student.objects.filter(teacher=teacher, status='منتظم')
        teacher_tested = ExamNomination.objects.filter(
            teacher=teacher,
            internal_passed=True
        ).count()
        
        # الغياب اليومي عند هذا المعلم
        teacher_absent_today = Attendance.objects.filter(
            student__teacher=teacher,
            date=today,
            status='غائب'
        ).count()

        # غياب المعلم نفسه اليوم
        teacher_self_absent_today = TeacherAttendance.objects.filter(
            teacher=teacher,
            date=today,
            status='غائب'
        ).count()
        
        # إحصائيات المعلم حسب المرحلة
        teacher_stage_stats = []
        for stage_code, stage_name in STAGES:
            stage_students = teacher_students.filter(educational_stage=stage_code)
            if stage_students.count() > 0:
                stage_tested = ExamNomination.objects.filter(
                    teacher=teacher,
                    student__educational_stage=stage_code,
                    internal_passed=True
                ).count()
                
                stage_absent = Attendance.objects.filter(
                    student__teacher=teacher,
                    student__educational_stage=stage_code,
                    date=today,
                    status='غائب'
                ).count()
                
                teacher_stage_stats.append({
                    'stage': stage_name,
                    'total_students': stage_students.count(),
                    'tested': stage_tested,
                    'absent_today': stage_absent,
                })
        
        teacher_stats.append({
            'teacher': teacher.username,
            'total_students': teacher_students.count(),
            'tested': teacher_tested,
            'absent_today': teacher_absent_today,
            'teacher_self_absent_today': teacher_self_absent_today,
            'stage_stats': teacher_stage_stats,
        })
    
    context = {
        'total_students': total_students,
        'total_tested': total_tested,
        'total_absent_today': total_absent_today,
        'total_teacher_absent_today': total_teacher_absent_today,
        'today': today,
        'stage_stats': stage_stats,
        'teacher_stats': teacher_stats,
    }
    
    return render(request, 'admin_statistics.html', context)

# 2. للمعلمين: أي مستخدم مسجل دخول (معلم أو مدير)
@login_required
def attendance_view(request):
    # كود التحضير
    return render(request, 'attendance.html')